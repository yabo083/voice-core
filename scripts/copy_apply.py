#!/usr/bin/env python3
"""Write a reviewed copy sheet back into the panel's source.

The sheet's identity is `file:line:column` plus the original text, and this trusts the text
over the position: a literal is replaced only where the original body is found, and a row
whose original is no longer there is reported rather than guessed at. Source has moved under
a sheet before (a trim pass between export and review), and a copy tool that silently writes
to the wrong offset would corrupt code while looking like it worked.

Usage:
    copy_apply.py <sheet.xlsx> [repo root] [--write]

Without `--write` it only reports, which is how you check a sheet before it touches anything.
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    write = "--write" in sys.argv
    sheet_path = Path(args[0])
    root = Path(args[1] if len(args) > 1 else ".")

    rows = list(load_workbook(sheet_path, data_only=True)["面板文案"].iter_rows(min_row=2, values_only=True))
    # file -> [(line, column, original, replacement)], newest edit per position wins.
    edits: dict[str, list[tuple[int, int, str, str]]] = defaultdict(list)
    unchanged = 0
    for file, line, column, _kind, original, new, _context in rows:
        if original is None:
            continue
        if new is None or str(new) == str(original):
            unchanged += 1
            continue
        edits[str(file)].append((int(line), int(column), str(original), str(new)))

    applied = skipped = 0
    problems: list[str] = []
    for file, items in sorted(edits.items()):
        path = root / file
        text = path.read_text(encoding="utf-8")
        lines = text.splitlines(keepends=True)
        # Bottom-up so an earlier replacement cannot move a later one's offset.
        for line_no, column, original, new in sorted(items, reverse=True):
            if line_no > len(lines):
                problems.append(f"{file}:{line_no} is past the end of the file")
                skipped += 1
                continue
            line = lines[line_no - 1]
            start = column  # 1-indexed column of the opening quote; body starts after it
            if line[start : start + len(original)] == original:
                at = start
            elif line.count(original) == 1:
                at = line.index(original)
            else:
                where = "not on that line" if original not in line else "ambiguous on that line"
                problems.append(f"{file}:{line_no} {where}: {original[:40]!r}")
                skipped += 1
                continue
            lines[line_no - 1] = line[:at] + new + line[at + len(original) :]
            applied += 1
        if write:
            path.write_text("".join(lines), encoding="utf-8")

    print(f"{'applied' if write else 'would apply'} {applied} edit(s) across {len(edits)} file(s)")
    print(f"unchanged rows: {unchanged}   skipped: {skipped}")
    for problem in problems:
        print(f"  ! {problem}")
    if problems and write:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
