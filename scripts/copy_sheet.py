#!/usr/bin/env python3
"""Collect every user-facing string in the panel into one spreadsheet for review.

Identity is `file:line:column` plus the original text, so an edited sheet maps back to
exactly one literal. The extractor takes any string literal containing CJK, which is what
"user-facing" means in this codebase: identifiers, class names, command names and event
names are ASCII by convention, and the few ASCII strings a user sees (units, `#rrggbb`
examples) are not worth reviewing.

Category comes from the syntax around the literal - `label:`, `title:`, `hint:`,
`placeholder=` and so on - because that is what decides whether a string is standing prose
or something only a hover shows, which is the distinction the review is about.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(sys.argv[1] if len(sys.argv) > 1 else ".")
OUT = Path(sys.argv[2] if len(sys.argv) > 2 else "panel-copy.xlsx")

CJK = re.compile(r"[\u3000-\u9fff\uff00-\uffef]")
# A string literal, single or double quoted or a backtick template. Non-greedy and NOT
# DOTALL: this runs per line, and a lone backtick in a template must not swallow the file.
LITERAL = re.compile(r"""(?P<q>['"`])(?P<body>(?:\\.|(?!\1).)*?)(?P=q)""")

# What the literal is doing, decided by the nearest syntax before it. A window rather than
# the line, because these object literals span lines: `el("p", {\n  class: "field__hint",\n
# text: "…" })` puts the class two lines above the string it describes.
WINDOW = 240

RULES: list[tuple[str, re.Pattern[str]]] = [
    # A class name is the most specific signal there is: it says where the text is painted.
    ("常驻说明", re.compile(r"""class:\s*["'`][^"'`]*(?:field__hint|form__hint|panel__hint|inv__hint)""")),
    ("日志", re.compile(r"""class:\s*["'`][^"'`]*console__""")),
    ("空状态", re.compile(r"emptyState\(|class:\s*[\"'`][^\"'`]*empty__")),
    ("屏幕标题", re.compile(r"""class:\s*["'`][^"'`]*screen__(?:title|desc)""")),
    ("预览文本", re.compile(r"""class:\s*["'`][^"'`]*preview__""")),
    # Then the call the literal is an argument of.
    ("hover 提示", re.compile(r"(?:withTip|blockedButton)\([^;]*$|(?:hint|reason|tip)\s*:\s*$")),
    ("提示框", re.compile(r"note\(\s*[\"'`](?:info|warn|fail|reuse)[\"'`]\s*,[^;]*$")),
    ("toast", re.compile(r"toast\(\s*(?:`|[\"'])?[^;]*$")),
    ("状态标签", re.compile(r"chip\([^;]*$")),
    ("按钮", re.compile(r"label\s*:\s*$|button\(\{[^}]*$")),
    ("面板标题", re.compile(r"(?:panel|expander)\(\{[^}]*title\s*:\s*$")),
    ("字段标签", re.compile(r"field\(\s*[\"'`][^\"'`]+[\"'`]\s*,\s*$")),
    ("占位符", re.compile(r"placeholder\s*:\s*$")),
    ("提示框标题", re.compile(r"title\s*:\s*$")),
    ("行文本", re.compile(r"text\s*:\s*$")),
]


# Checked only when nothing above matched, because these shapes carry no hint of their own
# purpose: a `Record` of labels and a bare array of strings look identical to any other
# literal, and the reviewer needs the context column to tell them apart.
FALLBACKS: list[tuple[str, re.Pattern[str]]] = [
    ("映射表", re.compile(r"(?m)^\s*(?:\[?[A-Za-z_$][\w$]*\]?|\"[^\"]+\")\s*:\s*$")),
    ("列表项", re.compile(r"(?:\[|,)\s*$")),
]


def classify(before: str) -> str:
    """`before` is the source up to the literal; the nearest rule that matches wins."""
    tail = before[-WINDOW:]
    best: tuple[int, str] | None = None
    for name, pattern in RULES:
        found = None
        for match in pattern.finditer(tail):
            found = match
        if found is not None and (best is None or found.start() > best[0]):
            best = (found.start(), name)
    if best is not None:
        return best[1]
    for name, pattern in FALLBACKS:
        if pattern.search(tail):
            return name
    return "其他"


def collect(path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    text = path.read_text(encoding="utf-8")
    starts = [0]
    for line in text.splitlines(keepends=True):
        starts.append(starts[-1] + len(line))
    for match in LITERAL.finditer(text):
        body = match.group("body")
        if not CJK.search(body):
            continue
        offset = match.start()
        number = max(i for i, start in enumerate(starts, start=1) if start <= offset)
        line = text[starts[number - 1] : starts[number] if number < len(starts) else len(text)]
        # A comment is for the next maintainer, not for the user.
        if line.lstrip().startswith(("//", "*", "/*")):
            continue
        rows.append(
            {
                "文件": path.as_posix(),
                "行": number,
                "列": offset - starts[number - 1] + 1,
                "类别": classify(text[:offset]),
                "原文": body,
                "新文（改这一列）": "",
                "上下文": line.strip()[:120],
            }
        )
    return rows


def main() -> None:
    targets = sorted(
        p
        for p in (ROOT / "manager" / "src").rglob("*.ts")
        if "node_modules" not in p.parts
    )
    rows: list[dict[str, object]] = []
    for path in targets:
        rows.extend(collect(path.relative_to(ROOT) if path.is_absolute() else path))

    book = Workbook()
    sheet = book.active
    sheet.title = "面板文案"
    headers = ["文件", "行", "列", "类别", "原文", "新文（改这一列）", "上下文"]
    sheet.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4A3A8C")
    for cell in sheet[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append([row[key] for key in headers])

    widths = {"文件": 34, "行": 6, "列": 6, "类别": 12, "原文": 60, "新文（改这一列）": 60, "上下文": 70}
    for index, key in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths[key]
    for column in ("E", "F", "G"):
        for cell in sheet[column]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{sheet.max_row}"

    # A second sheet the reviewer reads first: what each category means for how much text
    # is appropriate there, so the edits land in the right register.
    guide = book.create_sheet("怎么改")
    for line in [
        ("类别", "含在哪里出现", "写多少"),
        ("字段标签", "控件左边的名字", "一个词组，别写句子"),
        ("hover 提示", "只在鼠标悬停或键盘聚焦时出现", "一句，说机制/单位/文件；可以稍长"),
        ("按钮", "按钮上的字", "动词短语"),
        ("面板标题", "卡片标题", "名词，两到四个字"),
        ("占位符", "输入框里的灰字", "一个示例值，不是说明"),
        ("提示框", "常驻的 note", "只有三种情况该留：拦住用户的前置条件、破坏性操作的后果、失败原因+补救"),
        ("空状态", "什么都还没有时的那块", "可以两句：这里会出现什么、怎么让它出现"),
        ("状态标签", "chip 上的字", "两到四个字"),
        ("toast", "右下角一闪而过", "一句，说结果"),
        ("日志", "控制台里的行", "机器口吻，保留数字与路径"),
        ("其他", "分类器没认出来的", "看上下文列"),
        ("", "", ""),
        ("规则", "", ""),
        ("1", "控件自己就是解释：标签+值说完了它是什么，需要一句话的标签是标签起错了名", ""),
        ("2", "比标签长的东西放 hover，不放版面", ""),
        ("3", "不解释界面已经显示的事", ""),
        ("4", "数字和单位写进控件，不写在旁边的句子里", ""),
        ("5", "语气专业：说对象、单位、文件、机制；不写安抚，不写「会自己/就是/不用/也可以」", ""),
    ]:
        guide.append(list(line))
    guide.column_dimensions["A"].width = 14
    guide.column_dimensions["B"].width = 78
    guide.column_dimensions["C"].width = 46
    for cell in guide[1]:
        cell.font = head_font
        cell.fill = head_fill
    for column in ("B", "C"):
        for cell in guide[column]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    book.save(OUT)
    by_kind: dict[str, int] = {}
    for row in rows:
        by_kind[str(row["类别"])] = by_kind.get(str(row["类别"]), 0) + 1
    print(f"{len(rows)} strings from {len(targets)} files -> {OUT}")
    for kind, count in sorted(by_kind.items(), key=lambda item: -item[1]):
        print(f"  {kind:12} {count}")


if __name__ == "__main__":
    main()
